#!/usr/bin/env python3
"""
يعدّل ملف التقرير_اليومي_صندوق_الغدير.xlsx فقط:
- أوراق بأسماء شركات النظام (بما فيها الفرعية)
- شعارات في مكان LOGO
- ألوان مشتقة من لون كل شعار
يحافظ على هيكل التصميم الأصلي.
"""

from __future__ import annotations

import colorsys
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "التقرير_اليومي_صندوق_الغدير.xlsx"
PUBLIC = ROOT / "public"

# شركات النظام (بدون 010 التجريبي) — نفس ترتيب التقارير
COMPANIES = [
    {
        "sheet": "شركة الغدير",
        "title": "شركة الغدير",
        "logo": PUBLIC / "الغدير.png",
        "with_sample": True,
    },
    {
        "sheet": "شركة بدور بغداد",
        "title": "شركة بدور بغداد",
        "logo": PUBLIC / "بدور_بغداد.png",
        "with_sample": False,
    },
    {
        "sheet": "بدور بغداد - أمانات المستشار",
        "title": "بدور بغداد - صندوق امانات مصرف الستشار",
        "logo": PUBLIC / "بدور_بغداد.png",
        "with_sample": False,
    },
    {
        "sheet": "طيبة النجف",
        "title": "طيبة النجف",
        "logo": PUBLIC / "طيبة_النجف.png",
        "with_sample": False,
    },
    {
        "sheet": "غدير كربلاء",
        "title": "غدير كربلاء",
        "logo": PUBLIC / "غدير_كربلاء.png",
        "with_sample": False,
    },
    {
        "sheet": "بدور النجف",
        "title": "بدور النجف",
        "logo": PUBLIC / "بدور_النجف.png",
        "with_sample": False,
    },
    {
        "sheet": "الغدير - فرعي كربلاء",
        "title": "الغدير - صندوق فرعي - كربلاء",
        "logo": PUBLIC / "الغدير.png",
        "with_sample": False,
    },
    {
        "sheet": "غدير كربلاء - الفرعي",
        "title": "غدير كربلاء - الصندوق الفرعي",
        "logo": PUBLIC / "غدير_كربلاء.png",
        "with_sample": False,
    },
    {
        "sheet": "الغدير الفرعي - النجف",
        "title": "الغدير الفرعي - النجف",
        "logo": PUBLIC / "الغدير.png",
        "with_sample": False,
    },
]

# ألوان القالب الأصلية (للتعويض)
BASE = {
    "page": "F6F7FB",
    "white": "FFFFFF",
    "header": "CFE4F9",
    "sub": "D0E0F5",
    "soft": "E8F1FA",
    "logo_bg": "F0F6FC",
    "dark": "0F2E6B",
    "border": "A8C4E0",
    "border2": "7FA3C9",
    "green": "2E7D5B",
    "red": "C44B4B",
}


def hex_to_rgb(h: str):
    h = h.lstrip("#").upper()
    if len(h) == 8:
        h = h[2:]
    return tuple(int(h[i : i + 2], 16) for i in (0, 2, 4))


def rgb_to_hex(rgb):
    r, g, b = [max(0, min(255, int(x))) for x in rgb]
    return f"{r:02X}{g:02X}{b:02X}"


def mix(c1, c2, t):
    a = hex_to_rgb(c1)
    b = hex_to_rgb(c2)
    return rgb_to_hex(tuple(a[i] * (1 - t) + b[i] * t for i in range(3)))


def darken(h, amount=0.25):
    r, g, b = [x / 255 for x in hex_to_rgb(h)]
    hh, s, v = colorsys.rgb_to_hsv(r, g, b)
    v = max(0, v * (1 - amount))
    s = min(1, s * 1.1)
    return rgb_to_hex([x * 255 for x in colorsys.hsv_to_rgb(hh, s, v)])


def lighten(h, amount=0.7):
    return mix(h, "FFFFFF", amount)


# ألوان علامة تجارية ثابتة عند الحاجة (أدق من الاستخراج الآلي)
BRAND_PRIMARY = {
    "الغدير.png": "0070C0",
    "بدور_بغداد.png": "0D9A8C",
    "بدور_النجف.png": "15205A",
    "طيبة_النجف.png": "2F7FA3",
    "غدير_كربلاء.png": "9B163F",
}


def dominant_color(logo_path: Path) -> str:
    brand = BRAND_PRIMARY.get(logo_path.name)
    if brand:
        return brand

    im = PILImage.open(logo_path).convert("RGBA").resize((96, 96))
    counts = {}
    for r, g, b, a in im.getdata():
        if a < 180:
            continue
        if r > 245 and g > 245 and b > 245:
            continue
        # تجاهل الرمادي/البيج غير المشبع
        mx, mn = max(r, g, b), min(r, g, b)
        sat = (mx - mn) / (mx + 1)
        if sat < 0.18:
            continue
        key = (r // 12 * 12, g // 12 * 12, b // 12 * 12)
        counts[key] = counts.get(key, 0) + 1
    if not counts:
        return BASE["dark"]
    best = None
    best_score = -1
    for (r, g, b), n in counts.items():
        mx, mn = max(r, g, b), min(r, g, b)
        sat = (mx - mn) / (mx + 1)
        score = n * (0.25 + sat * 2.2)
        if score > best_score:
            best_score = score
            best = (r, g, b)
    return rgb_to_hex(best)


def theme_from_logo(logo_path: Path) -> dict:
    primary = dominant_color(logo_path)
    dark = darken(primary, 0.35)
    # لا تجعل النص غامقاً جداً حتى يصعب قراءته على الفاتح
    if sum(hex_to_rgb(dark)) < 80:
        dark = mix(dark, primary, 0.35)
    header = lighten(primary, 0.78)
    sub = lighten(primary, 0.72)
    soft = lighten(primary, 0.88)
    logo_bg = lighten(primary, 0.90)
    page = lighten(primary, 0.93)
    border = lighten(primary, 0.45)
    border2 = lighten(primary, 0.35)
    return {
        "page": page,
        "white": "FFFFFF",
        "header": header,
        "sub": sub,
        "soft": soft,
        "logo_bg": logo_bg,
        "dark": dark,
        "border": border,
        "border2": border2,
        "green": BASE["green"],
        "red": BASE["red"],
        "primary": primary,
    }


def cell_fill_rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    c = fill.fgColor
    if c.type == "rgb" and c.rgb:
        rgb = str(c.rgb).upper()
        if len(rgb) == 8:
            return rgb[2:]
        if len(rgb) == 6:
            return rgb
    if c.type == "theme":
        return None
    return None


def cell_font_rgb(cell) -> str | None:
    if not cell.font or not cell.font.color:
        return None
    c = cell.font.color
    if c.type == "rgb" and c.rgb:
        rgb = str(c.rgb).upper()
        if len(rgb) == 8:
            return rgb[2:]
        if len(rgb) == 6:
            return rgb
    return None


def map_color(old: str | None, theme: dict) -> str | None:
    if not old:
        return None
    old = old.upper()
    mapping = {
        BASE["page"]: theme["page"],
        BASE["header"]: theme["header"],
        BASE["sub"]: theme["sub"],
        BASE["soft"]: theme["soft"],
        BASE["logo_bg"]: theme["logo_bg"],
        BASE["dark"]: theme["dark"],
        BASE["border"]: theme["border"],
        BASE["border2"]: theme["border2"],
        "CFE4F9": theme["header"],
        "D0E0F5": theme["sub"],
        "E8F1FA": theme["soft"],
        "F0F6FC": theme["logo_bg"],
        "F6F7FB": theme["page"],
        "0F2E6B": theme["dark"],
        "A8C4E0": theme["border"],
        "7FA3C9": theme["border2"],
    }
    return mapping.get(old)


def recolor_sheet(ws, theme: dict):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 50, max_col=13):
        for cell in row:
            old_fill = cell_fill_rgb(cell)
            new_fill = map_color(old_fill, theme)
            if new_fill:
                cell.fill = PatternFill("solid", fgColor=new_fill)

            old_font = cell_font_rgb(cell)
            new_font = map_color(old_font, theme)
            if new_font:
                cell.font = Font(
                    name=cell.font.name or "Calibri",
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=new_font,
                )

            # حدود زرقاء → حدود الثيم
            if cell.border:
                def side_map(side: Side):
                    if not side or not side.color or not side.color.rgb:
                        return side
                    rgb = str(side.color.rgb).upper()
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    mapped = map_color(rgb, theme)
                    if not mapped:
                        return side
                    return Side(style=side.style, color=mapped)

                cell.border = Border(
                    left=side_map(cell.border.left),
                    right=side_map(cell.border.right),
                    top=side_map(cell.border.top),
                    bottom=side_map(cell.border.bottom),
                )


def clear_sample_rows(ws):
    """يفرغ صفوف البيانات التجريبية مع الإبقاء على التسلسل والصيغ."""
    for r in range(12, 23):
        for c in range(2, 14):  # B..M
            cell = ws.cell(r, c)
            # لا تمسح إن كانت صيغة
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None
    # صفوف ملخص العدد/المجموع التي فيها قيم ثابتة غير الصيغ تُترك للصيغ


def update_company_texts(ws, title: str):
    # العنوان المدمج D2
    ws["D2"].value = f"التقرير اليومي لصندوق {title}"
    # اسم الصندوق
    ws["C6"].value = title
    # مكان الشعار — اتركه فارغاً لأن الشعار صورة
    if ws["K2"].value and "شعار" in str(ws["K2"].value):
        ws["K2"].value = None


def remove_images(ws):
    # امسح كل الصور القديمة (أيقونات + placeholder)
    if hasattr(ws, "_images"):
        ws._images = []


def prepare_logo_png(src: Path, out: Path, max_w=220, max_h=140):
    im = PILImage.open(src).convert("RGBA")
    im.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
    # خلفية شفافة تبقى؛ إن كانت JPEG أصلاً محفوظة بامتداد png
    canvas = PILImage.new("RGBA", (max_w, max_h), (255, 255, 255, 0))
    x = (max_w - im.width) // 2
    y = (max_h - im.height) // 2
    canvas.paste(im, (x, y), im)
    out.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(out, "PNG")
    return out


def place_logo(ws, logo_path: Path, cache_dir: Path, company_key: str):
    prepared = prepare_logo_png(logo_path, cache_dir / f"{company_key}.png")
    img = XLImage(str(prepared))
    # موضع مكان الشعار K2:M8 — أعمدة K=11
    img.anchor = "K2"
    # أبعاد العرض تقريباً ضمن الصندوق
    img.width = 200
    img.height = 125
    ws.add_image(img)


def clear_data_area_values_keep_structure(ws):
    clear_sample_rows(ws)


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"missing {XLSX_PATH}")

    backup = XLSX_PATH.with_suffix(".xlsx.bak")
    shutil.copy2(XLSX_PATH, backup)

    wb = load_workbook(XLSX_PATH)
    # القالب: ورقة بيانات + ورقة فارغة
    sample_src = wb[wb.sheetnames[0]]
    empty_src = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else sample_src

    # احذف كل الأوراق القديمة بعد أخذ النسخ
    # ننشئ أوراقاً جديدة أولاً ثم نحذف القديمة
    created = []
    cache = ROOT / ".tmp_daily_report_logos"

    for i, company in enumerate(COMPANIES):
        src = sample_src if company["with_sample"] else empty_src
        ws = wb.copy_worksheet(src)
        # اسم مؤقت ثم نعيد التسمية بعد حذف التعارض
        temp_name = f"__tmp_{i}"
        ws.title = temp_name
        created.append((ws, company))

    # احذف الأوراق الأصلية (غير المؤقتة)
    for name in list(wb.sheetnames):
        if not name.startswith("__tmp_"):
            del wb[name]

    for ws, company in created:
        sheet_name = company["sheet"][:31]
        ws.title = sheet_name
        theme = theme_from_logo(company["logo"])
        remove_images(ws)
        recolor_sheet(ws, theme)
        update_company_texts(ws, company["title"])
        if not company["with_sample"]:
            clear_data_area_values_keep_structure(ws)
        place_logo(ws, company["logo"], cache, sheet_name.replace(" ", "_"))
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False

    wb.save(XLSX_PATH)
    print("saved", XLSX_PATH)
    print("sheets:", wb.sheetnames)
    print("backup:", backup)


if __name__ == "__main__":
    main()
